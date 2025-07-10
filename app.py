import io
from datetime import datetime
import streamlit as st
import pandas as pd
import chardet
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, numbers
from openpyxl.utils import get_column_letter

# Patrones de colores
# ! Secciones
# ? Indicaciones
# * Explicación
# todo: Detalles

# ! Configuraciones
st.set_page_config(page_title="CONDA web", layout="centered", page_icon=":snake:")

# ! Variables
date_format = "DD/MM/YYYY HH:MM"
lista_archivos = []
extensiones_excel = ["xl", "xlsx", "xls"]
limited_words = 50
# todo: Evitar repetición de guardados
list_process = []
# todo: Permitir la entrada únicamente de estos archivos
valid_extensions = ['xl', 'xlsx', 'csv']
# todo: Limite de archivos aceptables para consolidar
limited_files = 10

# ! Sessions
if "lista_archivos" not in st.session_state:
    st.session_state["lista_archivos"] = []

if "receive_files" not in st.session_state:
    st.session_state["receive_files"] = []

if "df_files" not in st.session_state:
    st.session_state["df_files"] = []

if "consolidate_df" not in st.session_state:
    st.session_state["consolidate_df"] = []

# ! Bandera
# todo: Bandera para activar globos
if "flag" not in st.session_state:
    st.session_state["flag"] = 0

# ! Functions
# todo: Guarda en memoria los resultados para no volver a ejecutar por cada llamada
# ?  Comprueba que no se repitan archivos
def validate_file():
    container = st.container(height=320)
    container.subheader(":material/assignment:  Archivos seleccionados", divider="red", help="Un archivo duplicado **NO** será tomado en cuenta para el proceso de consolidación... Bórrelo de la lista haciendo clic en X, o ignore los mensajes de advertencia :material/error:")
    # ? Ejecuta función si la condición se cumple
    # ? [test]Comprobar existencia de archivo para evitar repetición
    # ? [aprobado][25-06-25]
    # todo: Valida que los archivos subidos no sean repetidos para optimizar pprocesamiento
    archivo = st.session_state["archivos_subidos"]
    lista_nombres = []
    for id, archivo in enumerate(st.session_state["archivos_subidos"]):
        if archivo.name not in lista_nombres:
            file_name = archivo.name
            file_id = archivo.file_id
            file_content = archivo

            container.write(f"{id+1} - {file_name}")
            lista_consolidar = st.session_state["archivos_subidos"]
            lista_archivos.append(
                {"1": file_id, "nombre": file_name, "content": file_content}
            )
            lista_nombres.append(archivo.name)
        else:
            st.toast(f"{archivo.name} ya se encuentra **seleccionado**", icon=":material/error:")
        st.session_state["lista_archivos"] = lista_nombres
        st.session_state["receive_files"] = lista_archivos


# ? Extrae encode
# ? [test] Asegurar que los archivos en procesamiento sean los mismos que los subidos
# ? [aprobado] ! Ajustar
# ? [solución] Eliminar caché data
def extract_code(file_arg):
    # * [legacy] Recorre los archivos guardados en state
    # ? [test] Evluar extensión del archivo y convertir a df
    # ? [aprobado] Evluar extensión del archivo y convertir a df
    # ? [solución] Permite convertir archivos Excel y CSV, estos con código utf-8 y 16 a Dataframe sin corromperse en el proceso
    for id, archivo in enumerate(file_arg):
        # todo: Características del archivo
        file_name = archivo.name.split(".")[0]
        file_extension = archivo.name.split(".")[1]
        file_content = archivo
        if file_extension in extensiones_excel:
            read_file(
                archivo_nombre=file_name,
                archivo=file_content,
                extension=file_extension,
            )
        else:
            file_content = file_content.read()
            file_decode = chardet.detect(file_content[:10000])["encoding"] or "utf-8"
            if file_decode != "UTF-16":
                # todo: Agregar mensaje de asistencia en print
                read_file(file_name, file_content, file_decode, file_extension)
            elif file_decode == "UTF-16":
                read_file(file_name, file_content, file_decode, file_extension)


# ? Consolida los archivos
# todo: Solicita el diccionario del archivo
def read_file(archivo_nombre, archivo, encode=None, extension=None, aux=None):
    """
    Identifica qué extensión es el archivo.

    ### Args:
        file_name: Nombre del archivo extraído
        file_content: Contenido del archivo
        file_Encode: Código de archivo
        file_extension: Tipo de archivo
        aux: Archivo auxiliar
    """
    if extension in extensiones_excel:
        # todo: Agregar mensaje de asistencia en print
        df = pd.read_excel(archivo)

    elif extension == "csv" and encode == "UTF-16":
        # todo: Agregar mensaje de asistencia en print
        archivo = archivo.decode("utf-16")
        df = pd.read_csv(
            io.StringIO(archivo),
            sep=";",
            dtype=str,
            engine="python",
            skipinitialspace=True,
        )
        # ? Limpiar columnas
        df = limpiar_columnas(df)

    elif extension == "csv" and encode != "UTF-16":
        # todo: Agregar mensaje de asistencia en print
        # ? [testing] Convertir csv a df
        # ? [aprobado] Convertir csv a df
        file_buffer = io.BytesIO(archivo)
        df = pd.read_csv(file_buffer, encoding=encode)
        df = limpiar_columnas(df)
    st.session_state["df_files"].append(
        {"nombre": archivo_nombre, "extension": extension, "df": df}
    )


# ? Limpiar columnas de las tablas
def limpiar_columnas(df):
    df.columns = [col.encode("utf-8").decode("utf-8-sig").strip() for col in df.columns]
    return df


# ! Funciones para consolidar archivos
# ? [test] Motor de consolidación de datos
# ? [por hacer] Agregar comentarios de asistencia
def procesar_consolidacion(df_name, df):
    # * Extrae valores del archivo DataFrame
    df = df["df"]
    # * consolidar datos
    df_consolidado = consolidar_datos(df.copy())
    # * validar que se haya realizado correctamente
    if df_consolidado.empty:
        st.error("No se ha logrado este proceso")

    # * Conversión de fechas
    df_consolidado = convertir_fechas(df_consolidado)

    excel_file = crear_archivo_excel(df_consolidado)

    csv_file = crear_archivo_csv(df_consolidado)

    return {
        "file_name": df_name,
        "file_consolidate": df_consolidado,
        "file_excel": excel_file,
        "file_csv": csv_file,
    }


def consolidar_datos(df):
    # * Extrae primera columna
    f_column = df.columns[0]
    df_consolidate = df.copy()
    # * Recortar fechas en 16 carácteres
    df_consolidate[f_column] = df_consolidate[f_column].astype(str).str.slice(0, 16)
    # * Limpieza
    df_consolidate = df_consolidate[df_consolidate[f_column] != ""]
    # * Conversión a numérico
    col_num = []
    for col in df.columns[1:]:
        try:
            df_consolidate[col] = pd.to_numeric(df_consolidate[col], errors="coerce")
            col_num.append(col)
        except:
            pass

    # * Consolidación
    if col_num:
        df_consolidate = df_consolidate.groupby(f_column)[col_num].mean()
    else:
        df_consolidate = df_consolidate.groupby(f_column).first()

    return df_consolidate


def convertir_fechas(df):
    df_rest = df.copy()
    df_rest.index = pd.to_datetime(df_rest.index, errors="coerce")
    # * Elimina fechas inválidas
    return df_rest


def crear_archivo_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=True, sheet_name="Consolidado")

    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    date_style = NamedStyle(name="datetime", number_format="DD/MM/YYYY HH:MM")

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value:
            cell.style = date_style

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)

        if column_letter == "A":
            adjusted_width = max(adjusted_width, 20)

        ws.column_dimensions[column_letter].width = adjusted_width

    end_output = io.BytesIO()
    wb.save(end_output)
    end_output.seek(0)
    return end_output.getvalue()


def crear_archivo_csv(df, encode="cp1252"):
    file_csv = df.reset_index()
    fecha_col = file_csv.columns[0]
    file_csv[fecha_col] = file_csv[fecha_col].dt.strftime("%d/%m/%Y %H:%M")
    output = io.StringIO()
    file_csv.to_csv(output, index=False, sep=",")
    # ? [test | 02/07/25] Permite a los archivos CSV exportar con acentos
    # ? [APROBADO]
    return output.getvalue().encode("utf-8-sig")


# ? Almacenar en estado el resultado de la consolidación, exportación y resultado total del procesamiento del archivo
@st.cache_data
def save_files(df_files):
    # ? [test | 02/07/2025] Evaluar que no se consoliden dos veces el mismo archivo
    # ? [test | Aprobado]
    lista_consolidados = []
    for id, file_not_consolidate in enumerate(df_files):
        name_file = (
            f"{file_not_consolidate["nombre"]}.{file_not_consolidate["extension"]}"
        )
        if name_file not in lista_consolidados:
            lista_consolidados.append(file_not_consolidate)
        else:
            st.toast(f"El archivo {file_not_consolidate["nombre"]} está repetido")
    for id, df_file in enumerate(lista_consolidados):
        df_name = f"{df_file["nombre"]}.{df_file["extension"]}"
        for file_process in st.session_state["consolidate_df"]:
            file_process_name = f"{file_process["nombre"]}"
            list_process.append(file_process_name)
        if df_name not in list_process:
            rest = procesar_consolidacion(df_name, df_file)
            file_name = rest["file_name"]
            file_extension = df_file["extension"]
            file_content = rest["file_consolidate"]
            file_excel = rest["file_excel"]
            file_csv = rest["file_csv"]
            st.session_state["consolidate_df"].append(
                {
                    "nombre": file_name,
                    "settings_file_name": file_name,
                    "extension": file_extension,
                    "contenido": file_content,
                    "excel": file_excel,
                    "csv": file_csv,
                }
            )
        else:
            print("Archivo ya registrado")
        
    # ? -------------------------------------
    print("Archivos en memoria")
    st.session_state["flag"] = 1

@st.dialog("Editar nombre del archivo")
def settings(id, file_name):
    settings_file_name = st.text_input(
        f":material/edit_note: Agrega un **nuevo nombre** al archivo: {file_name}",
        max_chars=limited_words,
        placeholder="Nombre nuevo",
        value=file_name,
        type='default'
    )
    if st.button("Aplicar cambios", type="primary"):
        st.session_state["consolidate_df"][id]["settings_file_name"] = settings_file_name
        st.rerun()


# ! Skeleton Squeme
# ? Columnas
col_1, col_2 = st.columns(2, vertical_alignment="top", gap="small")

with col_1:
    st.title("Conda Web")
    archivos = st.file_uploader(
        "**Selecciona uno o varios archivos** haciendo clic aquí abajo 👇",
        accept_multiple_files=True,
        args="archivo",
        key="archivos_subidos",
        help=f"El sistema **solo** permite consolidar **{limited_files} archivos**...",
        type=valid_extensions
    )
    
    if archivos == []:
        st.caption(":material/info: Puedes subir **un archivo** de hasta 200 MB. No se considera el peso **total acumulado**...")

with col_2:
    # ? Evitar que se repitan archivos
    # todo: Revisa que haya información en la sesión y que la variable archivos tenga información
    if st.session_state["archivos_subidos"] != None and len(archivos) > 0:
        container = st.container(border=True)
        # ? Ejecuta función si la condición se cumple
        validate_file()
        consolidar = st.button(
            "**Consolidar datos** :material/upload_file:", type="primary", use_container_width=True, disabled=False,
            help="Haz clic sobre este botón para empezar a consolidar los **archivos seleccionados**"
        )
    
        if consolidar: 
            # ? Valida que el sistema no reciba más de una cantidad específica de archivos para su procesamiento en el servidor
            if len(archivos) > limited_files:
                st.error(f"El sistema solo permite consolidar **{limited_files} archivos**")
            else:
                # ? [test | 02/07/25] Evitar error al ingresar más archivos al sistema con archivos ya consolidados
                # ? [test] Aprobado | Procesa archivos nuevos con archivos subidos anteriormente
                lista_procesados = []
                lista_por_procesar = []

                if st.session_state["df_files"]:
                    for id, file_in in enumerate(st.session_state["df_files"]):
                        name_state_file = f"{st.session_state["df_files"][id]["nombre"]}.{st.session_state["df_files"][id]["extension"]}"
                        lista_procesados.append(name_state_file)

                for id, file_in in enumerate(archivos):
                    name_file = file_in.name
                    if name_file not in lista_procesados:
                        st.session_state["flag"] = 1
                        print(f"El archivo {name_file} no se ha procesado")
                        lista_por_procesar.append(file_in)

                # ? [test | 02/07/25] Mostrar toast de archivos
                if st.session_state["flag"] > 0:
                    # * Variables para asignación de mensajes
                    value = len(lista_por_procesar) if len(lista_por_procesar) > 1 else "un"
                    message = "archivos" if len(lista_por_procesar) >= 2 else "archivo" 
                    st.toast(f":material/access_time: Empezando el procesamiento de **{value}** {message}... Espere.")
                st.session_state["flag"] = 0
                extract_code(lista_por_procesar)
    # ? ---------------------------------------------#

    elif st.session_state["archivos_subidos"] != None and len(archivos) == 0:
        # todo: Muestra el mensaje de asistencia [SUJETOS A CAMBIOS]
        st.caption("Aquí se **mostrarán** cada uno los archivos que hayas **seleccionado**... ¡A consolidar datos!")

st.divider()

# ! Pie de página si no hay archivos en dataframes

if len(st.session_state["df_files"]) == 0 or len(archivos) == 0:
    # todo: Elimina la caché del estado en caso que se desee consolidar otros archivos

    save_files.clear()
    if "lista_archivos" in st.session_state:
        st.session_state["lista_archivos"] = []

    if "receive_files" in st.session_state:
        st.session_state["receive_files"] = []

    if "df_files" in st.session_state:
        st.session_state["df_files"] = []

    if "consolidate_df" in st.session_state:
        st.session_state["consolidate_df"] = []
    
    if "flag" in st.session_state:
        st.session_state["flag"] = 0
    
    
    # * Muestra mensaje de asistencia
    st.caption(
        "Aquí se mostrarán los **archivos** 📄 cuando hayan terminado de **consolidarse**. Cuando los selecciones, haz clic sobre el botón de arriba 👆 en **¡Consolidar datos!**"
    )
else:
    # ! Funciones
    save_files(st.session_state["df_files"])
    if st.session_state["flag"] > 0:
        st.balloons()
        st.session_state["flag"] = 0

    # ! Esqueleto de la página
    st.subheader("Archivos procesados", help="**Protip**: Al descargar el archivo... ¡Haz clic sobre el nombre en la tarjeta para que sea ocultado y no te confundas!")
    cols = st.columns(5, gap="small", vertical_alignment="top")

    # ? Determinar cuantas columnas mostrar
    cols_por_fila = 2

    # * Columnas padre
    for i in range(0, len(st.session_state["consolidate_df"]), cols_por_fila):
        cols = st.columns(cols_por_fila)
        # * Columnas anidadas
        for j, idx in enumerate(
            range(i, min(i + cols_por_fila, len(st.session_state["consolidate_df"])))
        ):
            file_in = st.session_state["consolidate_df"][idx]
            file_name = file_in["settings_file_name"].split(".")[0]
            file_name_expansive = f"{idx + 1} - " + (
                file_name
                if len(file_name) < 25
                else file_name[:-5] + "..."
            )
            extension = st.session_state["consolidate_df"][idx]["extension"].split(".")[
                -1
            ]
            # * Mostrar columnas
            with cols[j]:
                with st.expander(f"**{file_name_expansive}**", expanded=True):
                    st.write(f"Extensión del archivo: **{extension}**")
                    if st.button(
                        " :material/edit: Cambiar nombre",
                        use_container_width=True,
                        key=f"conf_{idx}",
                        type="secondary",
                    ):
                        settings(idx, file_name=file_name)
                    st.download_button(
                        ":material/file_download: Descargar como **Excel**",
                        data=file_in["excel"],
                        key=f"excel_{idx}",
                        file_name=f"Consolidado_{file_name}.xlsx",
                        use_container_width=True,
                    )
                    st.download_button(
                        " :material/file_download: Descargar como **CSV**",
                        data=file_in["csv"],
                        mime= "text/csv", 
                        key=f"csv_{idx}",
                        file_name=f"Consolidado_{file_name}.csv",
                        type="secondary",
                        help="Una vez abras el archivo, ¡ajuta la primera columna!",
                        use_container_width=True,
                    )
